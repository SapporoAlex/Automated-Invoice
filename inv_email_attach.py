import openpyxl as xl
from datetime import datetime
import time
from email.message import EmailMessage
import ssl
import smtplib


current_datetime = datetime.now()
current_year = current_datetime.year
current_month = current_datetime.month
current_day = current_datetime.day

# making sure an invoice created in early January,
# will identify the invoice year as the previous year
if current_month == 1 and current_day < 20:
    invoice_year = current_year - 1
else:
    invoice_year = current_year

# making sure an invoice creating in early January,
# will identify the invoice month as December,
# or if created in the last half of the month,
# the invoice month will be the current month
if current_month == 1 and current_day < 15:
    invoice_month = 12
elif current_day < 15:
    invoice_month = current_month - 1
elif current_day > 15:
    invoice_month = current_month


# determining following month
if invoice_month == 12:
    following_month = 1
else:
    following_month = int(invoice_month) + 1

# determining the last date of the month
months_of_31days = (1, 3, 5, 7, 8, 10, 12)
months_of_30days = (4, 6, 9, 11)
month_of_28days = 2
for m in months_of_30days:
    if invoice_month == m:
        invoice_date = 30
for m in months_of_31days:
    if invoice_month == m:
        invoice_date = 31
if invoice_month == 2:
    invoice_date = 28

# dictionary to help locate the correct Excel Sheet using the numeric month
calendar_dictionary = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December"
}
a = invoice_month
find_sheet = calendar_dictionary[a]

# loading local payment Excel file
payments_xl_file = xl.load_workbook(f"{invoice_year} Payments.xlsx")
payments_xl_file_sheet = payments_xl_file[f"{find_sheet}"]

# adding the amounts in the fee column until a cell doesn't have a fee
total_payment = 0
for row in range(2, payments_xl_file_sheet.max_row):
    cell = payments_xl_file_sheet.cell(row, 4)
    if cell.value is not None:
        total_payment += cell.value

# adding total to Payments Excel file
total_payment_cell = payments_xl_file_sheet["d21"]
total_payment_cell.value = total_payment
payments_xl_file.save(f"{invoice_year} Payments.xlsx")

# template
invoice_template = xl.load_workbook("202X年X月翻訳.xlsx")
invoice_sheet = invoice_template["請求書"]

# invoice cells to update
invoice_number_cell = invoice_sheet["g2"]
invoice_date_cell = invoice_sheet["g3"]
big_box_cell = invoice_sheet["c11"]
kingaku_cell = invoice_sheet["g14"]
gokei_cell = invoice_sheet["g31"]
item_name_cell = invoice_sheet["b14"]

# updated cells
invoice_number_cell.value = f"{invoice_year}{invoice_month}"
invoice_date_cell.value = f"{invoice_date}/{invoice_month}/{invoice_year}"
big_box_cell.value = f"{total_payment}円"
kingaku_cell.value = total_payment
gokei_cell.value = total_payment
item_name_cell.value = f"{invoice_year}年{invoice_month}月分翻訳料"

# saving invoice as a new Excel doc
invoice_template.save(f"{invoice_year}年{invoice_month}月翻訳.xlsx")

# writing and sending email
email_sender = '##########.com'
email_password = '###############'
email_receiver = '#######.com'

subject = "今月の請求書"
body = f"""
###、

いつもお世話になっております。
{invoice_month}月の請求書をお送りします。
遅くなり申し訳ございません。

{following_month}月の作業不可日はありません。

ご確認宜しくお願いいたします。

マッキンリー
"""

# Files to attach
payments_file = f"{invoice_year} Payments.xlsx"
invoice_file = f"{invoice_year}年{invoice_month}月翻訳.xlsx"

em = EmailMessage()
em['From'] = email_sender
em['To'] = email_receiver
em['Subject'] = subject
em.set_content(body)

with open(payments_file, "rb") as f:
    em.add_attachment(f.read(), filename=f"{invoice_year} Payments.xlsx", maintype="application",
                      subtype="octet-stream")

with open(invoice_file, "rb") as f:
    em.add_attachment(f.read(), filename=f"{invoice_year}年{invoice_month}月翻訳.xlsx",
                      maintype="application", subtype="octet-stream")

context = ssl.create_default_context()

with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
    smtp.login(email_sender, email_password)
    smtp.sendmail(email_sender, email_receiver, em.as_string())

# Statement upon completion
print(f"""
Invoice and email generated and sent
successfully using the following information.

sent invoice
Year: {invoice_year}
Month: {invoice_month}
Last day of month: {invoice_date}
Total amount to be paid: {total_payment}

Sent Payments
Month: {invoice_month}
Total: {total_payment}

Sent email
Month of invoice: {invoice_month}
Following month: {following_month}
""")
time.sleep(20)
